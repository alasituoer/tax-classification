# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html

from openpyxl import Workbook, load_workbook
import time


class DictFromWebsitesPipeline(object):

    def __init__(self):
        pass

    def open_spider(self, spider):
        self.path_to_write = "/mnt/hgfs/windows_desktop/classification_and_coding/" +\
                "data/dictionary_building/from_shopping_websites/raw_data_crawled/"
        self.filename_to_write = "dict_from_tb_" +\
                time.strftime("%Y%m%d", time.localtime()) + ".xlsx"
	# 从存储爬虫结果位置读取已爬取的文件
	# 如果读取不到说明(之前没有爬取)没有生成, 那么新建文件
	try:
	    self.wb = load_workbook(self.path_to_write + self.filename_to_write)
	except Exception, e:
	    self.wb = Workbook()

    def process_item(self, item, spider):
        # 根据item['sheetname']确定应该将其存入excel文件的哪张表中
	# 如果该表已存在则覆盖, 如果不存在则新建
        try:
            self.ws = self.wb[item['sheetname']]
        except Exception,e:
            self.ws = self.wb.create_sheet(title=item['sheetname'])

        # 安装顺序写入item的指定各值(首页/idx/c1/c2/kw), 如果某值缺失则跳过
        list_to_write = []
        for k in ['idx', 'c1', 'c2', 'kw',]:
            try:
                list_to_write.append(item[k])
            except Exception, e:
                continue
#        print repr(list_to_write).decode('unicode-escape')
        self.ws.append(list_to_write)

        return item

    def close_spider(self, spider):
        self.wb.save(self.path_to_write + self.filename_to_write)



