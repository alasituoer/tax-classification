#coding:utf-8
import scrapy
import time
from openpyxl import Workbook, load_workbook

RENDER_HTML_URL = "http://jxdmyx.com:8050/render.html"

class Xdb1688Spider(scrapy.Spider):
    name = "xdb1688_spider"
    allowed_domains = ['1688.com', 'jxdmyx.com',]
    path_to_write = "/mnt/hgfs/windows_desktop/classification_and_coding/" +\
            "data/dictionary_building/from_shopping_websites/" +\
            "dict_from_xdb1688_" + time.strftime("%Y%m%d", time.localtime()) + ".xlsx"

    def __init__(self):
        try:
            wb = load_workbook(self.path_to_write)
        except Exception, e:
            wb = Workbook()
            wb.save(self.path_to_write)

    def start_requests(self):
        list_urls = ['https://xdb.1688.com/all-shop.html',]
	for url in list_urls:
	    yield scrapy.Request(url=url, callback=self.parse)
    
    def parse(self, response):
	dict_cate_url = {}
	for sel in response.xpath("//ul[@class='shops-collect']/li"):
            c1_name = sel.xpath("div/text()").extract()[0]
	    for sel2 in sel.xpath("ul/li"):
		c2_name = sel2.xpath("a/div[2]/h3/text()").extract()[0]
		c2_url = sel2.xpath("a/@href").extract()[0]
		dict_cate_url[c2_name] = c2_url
#	print dict_cate_url

        list_cate_tobe_crawled = [u'五金店',]
        list_methods = [self.crawlingWujin,]
        dict_cate_method = dict(zip(list_cate_tobe_crawled, list_methods))

        for cate_name in list_cate_tobe_crawled[:1]:
            cate_url = dict_cate_url[cate_name]
	    cate_url = RENDER_HTML_URL + "?url=" + cate_url + "&timeout=10&wait=2"
#	    print cate_name, cate_url
            yield scrapy.Request(url=cate_url, callback=dict_cate_method[cate_name])



    def crawlingWujin(self, response):
        """五金店"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'五金店'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'五金店')

	for sel in response.xpath("//ul[@class='xdb-goods-shelf-list fd-clr']/li"):
	    cate_name = sel.xpath("a/div/div[2]/text()").extract()[0]
	    localid = sel.xpath("@data-spm-click").extract()[0]
	    shelfResourceTags = localid.split(';')[-2]
	    print cate_name, shelfResourceTags[10:]
	    print "无法正确提取小店宝的品类下关键字"

