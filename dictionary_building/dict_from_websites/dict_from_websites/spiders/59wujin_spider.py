#coding:utf-8
import scrapy
import time
from openpyxl import Workbook, load_workbook

class ChinawujinSpider(scrapy.Spider):
    name = "59wujin_spider"
    allowed_domains = ['59wujin.com', 'jxdmyx.com',]
    path_to_write = "/mnt/hgfs/windows_desktop/classification_and_coding/" +\
            "data/dictionary_building/from_shopping_websites/" +\
            "dict_from_59wujin_" + time.strftime("%Y%m%d", time.localtime()) + ".csv"

#    def __init__(self):
#        try:
#            wb = load_workbook(self.path_to_write)
#        except Exception, e:
#            wb = Workbook()
#            wb.save(self.path_to_write)

    def start_requests(self):
        list_urls = ['http://www.59wujin.com/sellfl.html',]
	for url in list_urls:
	    yield scrapy.Request(url=url, callback=self.parse)
    
    def parse(self, response):
        dict_c1_url = {}
        for sel in response.xpath("//div[@class='wrap']/div[@class='pBox']"):
            c1_name = sel.xpath("h3/text()").extract()[0]
	    dict_c2_url = {}
            for sel2 in sel.xpath("ul/li"):
                c2_name = sel2.xpath("a/text()").extract()[0]
                c2_url = 'http://www.59wujin.com/' + sel2.xpath("a/@href").extract()[0]
#                print c1_name, c2_name, c2_url
                dict_c2_url[c2_name] = c2_url
#            print repr(dict_c2_url).decode('unicode-escape')
            dict_c1_url[c1_name] = dict_c2_url
#        print repr(dict_c1_url).decode('unicode-escape')
#	print repr(dict_c1_url.keys()).decode('unicode-escape')

	list_tobe_crawled = [u'电子元器件', u'机械及行业设备', u'五金、工具',]
	for idx,c1_name in enumerate(list_tobe_crawled):
	    for c2_name in dict_c1_url[c1_name].keys():
		c2_url = dict_c1_url[c1_name][c2_name]
                meta = {'idx': idx, 'c1_name': c1_name, 'c2_name': c2_name,}
#                print c1_name, c2_name, c2_url
                yield scrapy.Request(url=c2_url, 
                        callback=self.getListOfKeywords, meta=meta)

    def getListOfKeywords(self, response):
	with open(self.path_to_write, 'a') as f1:
            for sel in response.xpath("//div[@class='pBox']"):
                list_kws_c2 = sel.xpath("ul/li/a/text()").extract()
                for kw in list_kws_c2:
                    list_to_write = [str(response.meta['idx']+1), 
			    response.meta['c1_name'], response.meta['c2_name'], kw,]
#                    print ','.join(list_to_write)
                    f1.write(','.join(list_to_write) + "\n")




