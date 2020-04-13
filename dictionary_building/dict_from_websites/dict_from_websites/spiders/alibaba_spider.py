#coding:utf-8
import scrapy
import time
from openpyxl import Workbook, load_workbook

class AlibabaSpider(scrapy.Spider):
    name = "alibaba_spider"
    allowed_domains = ['1688.com']
    path_to_write = "/mnt/hgfs/windows_desktop/classification_and_coding/" +\
            "data/dictionary_building/from_shopping_websites/" +\
            "dict_from_alibaba_" + time.strftime("%Y%m%d", time.localtime()) + ".xlsx"

    def __init__(self):
        try:
            wb = load_workbook(self.path_to_write)
        except Exception, e:
            wb = Workbook()
            wb.save(self.path_to_write)

    def start_requests(self):
        list_urls = ['https://www.1688.com',]
        for url in list_urls:
            yield scrapy.Request(url=url, callback=self.parse)

    def parse(self, response):
        dict_cate_url = {}
        for sel in response.xpath("//ul[@id='sub-nav']/li/a"):
            cate_name = sel.xpath("text()").extract()[0]
            cate_url = sel.xpath("@href").extract()[0]
            if 'http' not in cate_url:
                cate_url = "https:" + sel.xpath("@href").extract()[0]
            dict_cate_url[cate_name] = cate_url
#        print dict_cate_url

        list_cate_tobe_crawled = [u'机械', u'五金工具',]
        list_methods = [self.crawlingJd, self.crawlingWjgj]
        dict_cate_method = dict(zip(list_cate_tobe_crawled, list_methods))

        for cate_name in list_cate_tobe_crawled:
            cate_url = dict_cate_url[cate_name]
#            print cate_name, cate_url
            yield scrapy.Request(cate_url, callback=dict_cate_method[cate_name])


    def crawlingWjgj(self, response):
        """五金工具"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'五金工具'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'五金工具')

        for idx,sel in enumerate(response.xpath("//div[@class='ch-menu-body']/div")):
            c1 = sel.xpath("h3/span/text()").extract()[0]
            list_kws_c1 = sel.xpath("div/ul/li/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, kw]
#                print repr(list_to_write).decode("unicode-escape")
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingJd(self, response):
        """机械"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'机械'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'机械')

        # 左侧导航栏
        list_c1 = []
        for idx,sel in enumerate(response.xpath(
                "//div[@class='ch-menu-body']/div[@class='ch-menu-item']")):
            c1 = sel.xpath("h3/span/text()").extract()[0]
            list_c1.append(c1)
            list_kws_c1 = sel.xpath("div/ul/li/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, '', kw,]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

        # 悬浮隐藏框
        for idx,sel in enumerate(response.xpath(
                "//div[@class='ch-menu-layer']/div/div")):
            for sel2 in sel.xpath("div"):
                c2 = sel2.xpath("h4/text()").extract()[0]
                list_kws_c2 = sel2.xpath("div/ul/li/a/text()").extract()
                for kw in list_kws_c2:
                    list_to_write = [idx+1, list_c1[idx], c2, kw,]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


