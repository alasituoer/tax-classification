#coding:utf-8
import scrapy
import numpy as np
import time
import json
from openpyxl import Workbook, load_workbook


RENDER_HTML_URL = "http://localhost:8050/render.html"

class TaobaoSpider(scrapy.Spider):
    name = 'taobao_spider'
    allowed_domains = ['alicdn.com', 'taobao.com', 'localhost', 'jiyoujia.com',]

    path_to_write = '/mnt/hgfs/windows_desktop/classification_and_coding/' +\
            'data/dictionary_building/from_shopping_websites/dict_from_taobao_' +\
            time.strftime("%Y%m%d", time.localtime()) + '.xlsx'

    def __init__(self):
        try:
            wb = load_workbook(self.path_to_write)
        except Exception, e:
            wb = Workbook()
            wb.save(self.path_to_write)

    def start_requests(self):
        #1 首页的浮动层类别及关键字
        for u in ['https://tce.alicdn.com/api/data.htm?ids=222887%2C222890%2C222889%2C222886%2C222906%2C222898%2C222907%2C222885%2C222895%2C222878%2C222908%2C222879%2C222893%2C222896%2C222918%2C222917%2C222888%2C222902%2C222880%2C222913%2C222910%2C222882%2C222883%2C222921%2C222899%2C222905%2C222881%2C222911%2C222894%2C222920%2C222914%2C222877%2C222919%2C222915%2C222922%2C222884%2C222912%2C222892%2C222900%2C222923%2C222909%2C222897%2C222891%2C222903%2C222901%2C222904%2C222916%2C222924']:
            yield scrapy.Request(url=u, callback=self.crawlingHeadfloat)

        #2 首页第一屏左侧品类的链接(子)页面
#        list_urls = ['https://www.taobao.com',]
#        for url in list_urls:
#            yield scrapy.Request(url=url, callback=self.parse)


    def crawlingHeadfloat(self, response):
        """首页浮动展示类别及关键字"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'首页全类别'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'首页全类别')

        json_data = json.loads(response.body)
        #for idx,cate_index in enumerate(json_data.keys()[23:24]):
        for idx,cate_index in enumerate(json_data.keys()):
            json_index = json_data[cate_index]['value']
            # 由于字典某些键值可能为空, 遇空则跳过
            try:
                for kw in json_index['list']:
                    list_to_write = [idx+1, json_index['head'][0]['name'], kw['name']]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
            except Exception, e:
                continue
        wb.save(self.path_to_write)



    def parse(self, response):
        """从首页第一屏的左侧品类, 得到其品类链接, 传递给指定的字页面函数解析"""
        for sel in response.xpath("//ul[@class='service-bd']"):
            list_cate_name = sel.xpath("li/a/text()").extract()
            list_cate_url = sel.xpath("li/a/@href").extract()

        # 爬取到的部分链接缺少'https:', 如: "//mei.taobao.com"
        list_cate_url = ['https:'+u if 'http' not in u else u for u in list_cate_url]

        dict_cate_url = dict(zip(list_cate_name, list_cate_url))
        #print len(list_cate_name)

        # 百货的链接指向阿里旗下的一站式家居购物平台jiyoujia
        # 且访问过多会被要求登陆, 不如直接访问"极有家"
        dict_cate_url[u'百货'] = 'https://www.jiyoujia.com'

        list_cate = [# 5个一行 21个大品类
                u'女装', u'男装', u'内衣', u'鞋靴', u'箱包',
                u'配件', u'童装玩具', u'家电', u'美妆', u'洗护',
                u'珠宝', u'眼镜', u'手表', u'运动', u'乐器',
                u'美食', u'汽车', u'办公', u'DIY', u'五金电子',
                u'百货',]
        list_method = [self.crawlingNvzhuang, self.crawlingNanzhuang,
                self.crawlingNeiyi, self.crawlingXie, self.crawlingXiangbao,
                self.crawlingPei, self.crawlingQbb, self.crawlingTbdc,
                self.crawlingMei, self.crawlingXihuyongpin, self.crawlingZhubao,
                self.crawlingYanjing, self.crawlingShoubiao, self.crawlingCoolcityhome,
                self.crawlingAmusement, self.crawlingChi, self.crawlingCar,
                self.crawlingBangong, self.crawlingDingzhi, self.crawlingWujin,
                self.crawlingBaihuo,]
        dict_cate_method = dict(zip(list_cate, list_method))

        # 测试时指定一个品类或只取最后一个
#        for cate_name in [u'内衣']:
#        for cate_name in list_cate[-1:]:
        # 正式爬取时则遍历
        for cate_name in list_cate:
            cate_url = dict_cate_url[cate_name]
#            print cate_name, cate_url
            cate_url = RENDER_HTML_URL + "?url=" + cate_url + "&timeout=10&wait=2"
            yield scrapy.Request(url=cate_url, callback=dict_cate_method[cate_name])


    def crawlingBaihuo(self, response):
        """百货"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'百货'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'百货')

        # 得到一级分类名
        for sel in response.xpath("//ul[@class='sub-menus clearfix']"):
            list_c1 = sel.xpath("li/a/text()").extract()
#            print repr(list_c1).decode('unicode-escape')

        # 4个大品类的二级分类名及关键字存放在<textarea>中
        # 但提取到的<textarea>标签的文本值是"一段html语句"(不是往常的'字典子字符串')
        # 所以重新指定response.body, 重新解析网页得到所需内容
        for idx,sel in enumerate(response.xpath(
                "//textarea[@class='J_Sub_Menu_Content']")):
            res2 = response.replace(body=sel.xpath("text()").extract()[0])
            # 显式复制
            for sel2 in res2.xpath("//div[@class='col-block-wrap']"):
                #c2 = sel2.xpath("a/img/@alt").extract()[0]
                c2 = sel2.xpath("a/text()").extract()[0]
                for sel3 in sel2.xpath("ul[@class='server-list clearfix']/li"):
                    list_kws_c2 =  sel3.xpath("a/text()").extract()[0]
                    list_to_write = [idx+1, list_c1[idx], c2, list_kws_c2]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingWujin(self, response):
        """五金电子"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'五金电子'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'五金电子')

        # 第一屏左侧分类(7个, 少于下述的9个
        # 所以不能在此得到list_c1做下述的大类别名匹配)
        for sel in response.xpath("//div[@class='sub-d-menu']/dl"):
            c1 = sel.xpath("dt/text()").extract()[0]
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = ['', c1, '', kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

        # 往下滚动左侧细分类
        # 左侧的(9个)类别名是图片, 需要单独提取
        list_c1 = []
        for sel in response.xpath("//p[@class='designer']"):
            c1 = sel.xpath("span/text()").extract()[0]
            list_c1.append(c1)
#        print repr(list_c1).decode('unicode-escape')
        # 再提取关键字
        for idx,sel in enumerate(response.xpath("//textarea[@class='J_Dynamic_Data']")):
            dict_old = eval(sel.xpath("text()").extract()[0].strip())
            list_needed = dict_old['textlink'][0]['text']
#            print json.dumps(list_needed, indent=1)
            for c2 in list_needed:
                for kw in c2['texts']:
                    list_to_write = [idx+1, list_c1[idx], c2['name'].decode('utf-8'), 
                            kw['cat_name'].decode('utf-8')]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingDingzhi(self, response):
        """DIY (定制)"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'DIY'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'DIY')

        #print 'hello alas'
        for sel in response.xpath("//div[@id='guid-361957']/div/textarea"):
            dict_old = eval(sel.xpath("text()").extract()[0])
            #print json.dumps(dict_old['cat_mian'], indent=1)
            for idx,c1 in enumerate(dict_old['cat_mian']):
                for c2 in c1['cat_data']:
                    list_to_write = [idx+1, c1['name'].decode('utf-8'),
                            c2['cat_name'].decode('utf-8')]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingBangong(self, response):
        """办公 (淘宝办公)"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'办公'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'办公')

        list_c1 = []
        for idx,sel in enumerate(response.xpath("//div[@class='sub-d-menu']/dl")):
            c1 = sel.xpath("dt/text()").extract()[0]
            list_c1.append(c1)
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, '', kw,]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

        for idx,sel in enumerate(response.xpath("//textarea[@class='J_Dynamic_Data']")):
            dict_old = eval(sel.xpath("text()").extract()[0].strip())
            list_needed = dict_old['textlink'][0]['text']
#            print json.dumps(list_needed, indent=1)
            for c2 in list_needed:
                for kw in c2['texts']:
                    list_to_write = [idx+1, list_c1[idx],
                            c2['name'].decode('utf-8'), kw['cat_name'].decode('utf-8')]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingCar(self, response):
        """汽车 (阿里汽车)"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'汽车'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'汽车')

        for idx,sel in enumerate(response.xpath("//div[@class='jia-left-nav']/ul/li")):
            c1 = sel.xpath("div/text()").extract()[0].strip()
            list_kws_c1 = sel.xpath("a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, '', kw,]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

            for sel2 in sel.xpath("div/div"):
                c2 = sel2.xpath("div/text()").extract()[0].strip()
                list_kws_c2 = sel2.xpath("div/a/text()").extract()
                for kw in list_kws_c2:
                    list_to_write = [idx+1, c1, c2.decode('utf-8'), kw.decode('utf-8'),]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingChi(self, response):
        """'美食','生鲜','零食', 淘宝汇吃"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'美食'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'美食')

        for idx,sel in enumerate(response.xpath(
                "//ul[@class='items-menu-container J_Nav']/li")):
            c1 = sel.xpath("div[@class='keys-wrap']/a/text()").extract()[0]
            list_kws_c1 = sel.xpath("div[@class='keys-wrap']/p/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, '', kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

            for sel2 in sel.xpath("div[@class='subitems-container']/ul/li"):
                c2 = sel2.xpath("a/span/text()").extract()[0]
                list_kws_c2 = sel2.xpath("div/span/a/text()").extract()
                for kw in list_kws_c2:
                    list_to_write = [idx+1, c1, c2, kw]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingAmusement(self, response):
        """乐器"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'乐器'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'乐器')

        for sel in response.xpath("//div[@id='guid-770244']/div"):
            dict_c1 = eval(sel.xpath("textarea/text()").extract()[0])
            for idx,c1 in enumerate(dict_c1['cat_mian']):
                for kw in c1['cat_data']:
                    list_to_write = [idx+1, c1['name'].decode('utf-8'),
                            kw['cat_name'].decode('utf-8')]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingCoolcityhome(self, response):
        """运动、户外 淘宝字页面称之为 酷玩城"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'运动'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'运动')

        for sel in response.xpath("//ul[@class='list-name']"):
            list_c1 = sel.xpath("li/a/text()").extract()
            list_c1 = [c.strip() for c in list_c1 if c.strip() != u'']
#            print repr(list_c1).decode('unicode-escape')

        for idx,sel in enumerate(response.xpath("//div[@class='list-k']/ul")):
            list_kws_c1 = sel.xpath("li/a/text()").extract()
            list_kws_c1 = [c.strip() for c in list_kws_c1]
#            print repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [idx+1, list_c1[idx], kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingShoubiao(self, response):
        """手表"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'手表'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'手表')

        for idx,sel in enumerate(response.xpath("//div[@class='bd']/ul/li")):
            c1 = sel.xpath("p/a/text()").extract()[0]
            list_kws_c1 = eval(sel.xpath("dl/textarea/text()").extract()[0])['custom']
            list_kws_c1 = [d['cat_name'].decode('utf-8') for d in list_kws_c1]
#            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, kw,]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingYanjing(self, response):
        """眼镜"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'眼镜'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'眼镜')

        for idx,sel in enumerate(response.xpath("//div[@class='bd']/ul/li")):
            c1 = sel.xpath("p/a/text()").extract()[0]
            list_kws_c1 = eval(sel.xpath("dl/textarea/text()").extract()[0])['custom']
            list_kws_c1 = [d['cat_name'].decode('utf-8') for d in list_kws_c1]
#            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, kw,]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingZhubao(self, response):
        """珠宝"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'珠宝'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'珠宝')

        for idx,sel in enumerate(response.xpath("//div[@class='bd']/ul/li")):
            c1 = sel.xpath("p/a/text()").extract()[0]
            list_kws_c1 = eval(sel.xpath("dl/textarea/text()").extract()[0])['custom']
            list_kws_c1 = [d['cat_name'].decode('utf-8') for d in list_kws_c1]
#            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, kw,]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingXihuyongpin(self, response):
        """洗护"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'洗护'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'洗护')

        list_c1 = []
        for sel in response.xpath("//div[@class='head-line']"):
            c1 = sel.xpath("h4/text()").extract()[0]
            list_c1.append(c1)
#        print repr(list_c1).decode('unicode-escape')

        for idx,sel in enumerate(response.xpath("//div[@class='list-wrap']")):
            list_kws_c1 = sel.xpath("p/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [idx+1, list_c1[idx], kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingMei(self, response):
        """美妆"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'美妆'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'美妆')

        for idx,sel in enumerate(response.xpath(
                "//div[@class='market-wrap clearfix sm-cat-list-main']/dl")):
            # 首页显示类别
            list_old_cate = eval(sel.xpath("textarea[1]/text()").extract()[0].strip())
            list_old_cate = [d['cat_name'].decode('unicode-escape') for d in list_old_cate]
#            print repr(list_old_cate).decode('unicode-escape')
            for kw in list_old_cate[1:]:
                list_to_write = [idx+1, list_old_cate[0], '', kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

            # 扩展(及隐藏)类别
            list_extra_cate = eval(sel.xpath("textarea[2]/text()").extract()[0].strip())
            list_cate = [d['cat_name'].decode('unicode-escape') for d in list_extra_cate]
            list_istitle = [d['is_title']=='true' for d in list_extra_cate]
#            print repr(list_cate).decode('unicode-escape')
#            print list_istitle
            list_index = np.argwhere(np.array(list_istitle)).T[0].tolist()
            list_index.append(len(list_cate))
            list_zip_cate_kw = [(list_index[i], list_index[i+1]) \
                    for i in range(len(list_index)-1)]
            for i,j in list_zip_cate_kw:
                for k in range(i+1,j):
                    list_to_write = [idx+1, list_old_cate[0], list_cate[i], list_cate[k]]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingTbdc(self, response):
        """家电、数码、手机 淘宝字页面称之为 淘宝电场"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'家电'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'家电')

        list_c1 = []
        for sel in response.xpath("//div[@class='nav-p']/p"):
            c1 = sel.xpath("span/text()").extract()[0]
            list_c1.append(c1)
#        print repr(list_c1).decode('unicode-escape')

        # 抓取二级类别及其关键字
        for i,sel in enumerate(response.xpath("//div[@class='nav-text']/div")):
            for sel2 in sel.xpath("div"):
                c2 = sel2.xpath("div/text()").extract()[0]
                list_kws_c2 = sel2.xpath("div/a/text()").extract()
                for kw in list_kws_c2:
                    list_to_write = [i+1, list_c1[i], c2, kw]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingQbb(self, response):
        """童装玩具(淘宝详细页标题 亲宝贝)"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'童装玩具'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'童装玩具')

        # 将返回存入本地文件, 查看网页是否是JS渲染之后的
#        with open('data/test.html', 'w') as f:
#            f.write(response.body)

        for i,sel in enumerate(response.xpath("//ul[@class='nav-lists']/li")):
            list_c1_and_kws = sel.xpath("a/text()").extract() 
#            print repr(list_c1_and_kws).decode('unicode-escape')
            for kw in list_c1_and_kws[1:]:
                list_to_write = [i+1, list_c1_and_kws[0], kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingPei(self, response):
        """配件, 与箱包的解析方式一致, 即下面只是更改了写入excel的表名"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'配件'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'配件')

        for idx,sel in enumerate(response.xpath("//dl[@class='theme-bd-level2']")):
            # 首页显示类别
            list_old_cate = eval(sel.xpath("textarea[1]/text()").extract()[0].strip())
            list_old_cate = [c['cat_name'].decode('utf-8') for c in list_old_cate]
#            print repr(list_old_cate).decode('unicode-escape')
            for kw in list_old_cate[1:]:
                list_to_write = [idx+1, list_old_cate[0], '', kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

            # 扩展(及隐藏)类别
            list_extra_cate = eval(sel.xpath("textarea[2]/text()").extract()[0].strip())
            list_cate = [d['cat_name'].decode('utf-8') for d in list_extra_cate]
            list_istitle = [d['is_title']=='true' for d in list_extra_cate]
#            print repr(list_cate).decode('unicode-escape')
#            print list_istitle
            list_index = np.argwhere(np.array(list_istitle)).T[0].tolist()
            list_index.append(len(list_cate))
            list_zip_cate_kw = [(list_index[i], list_index[i+1]) \
                    for i in range(len(list_index)-1)]
            for i,j in list_zip_cate_kw:
                for k in range(i+1,j):
                    list_to_write = [idx+1, list_old_cate[0], list_cate[i], list_cate[k]]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingXiangbao(self, response):
        """箱包"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'箱包'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'箱包')

        for idx,sel in enumerate(response.xpath("//dl[@class='theme-bd-level2']")):
            # 首页显示类别
            list_old_cate = eval(sel.xpath("textarea[1]/text()").extract()[0].strip())
            list_old_cate = [c['cat_name'].decode('utf-8') for c in list_old_cate]
#            print repr(list_old_cate).decode('unicode-escape')
            for kw in list_old_cate[1:]:
                list_to_write = [idx+1, list_old_cate[0], '', kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

            # 扩展(及隐藏)类别
            list_extra_cate = eval(sel.xpath("textarea[2]/text()").extract()[0].strip())
            list_cate = [d['cat_name'].decode('utf-8') for d in list_extra_cate]
            list_istitle = [d['is_title']=='true' for d in list_extra_cate]
#            print repr(list_cate).decode('unicode-escape')
#            print list_istitle
            list_index = np.argwhere(np.array(list_istitle)).T[0].tolist()
            list_index.append(len(list_cate))
            list_zip_cate_kw = [(list_index[i], list_index[i+1]) \
                    for i in range(len(list_index)-1)]
            for i,j in list_zip_cate_kw:
                for k in range(i+1,j):
                    list_to_write = [idx+1, list_old_cate[0], list_cate[i], list_cate[k]]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingXie(self, response):
        """鞋靴"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'鞋靴'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'鞋靴')

        for i,sel in enumerate(response.xpath("//dl[@class='theme-bd-level2']")):
            list_old_cate = eval(sel.xpath("textarea[1]/text()").extract()[0].strip())
            list_old_cate = [c['cat_name'].decode('utf-8') for c in list_old_cate]
#            print repr(list_old_cate).decode('unicode-escape')
            for kw in list_old_cate[1:]:
#                print list_old_cate[0], kw
                ws.append([2*i+1, list_old_cate[0], kw])

            # 每一大类的扩展类单独作为一个大类
            list_extra_cate = eval(sel.xpath("textarea[2]/text()").extract()[0].strip())
            list_extra_cate = [c['cat_name'].decode('utf-8') for c in list_extra_cate]
#            print repr(list_extra_cate).decode('unicode-escape')
            for kw in list_extra_cate[1:]:
#                print list_extra_cate[0], kw
                ws.append([2*i+2, list_extra_cate[0], kw])
        wb.save(self.path_to_write)


    def crawlingNeiyi(self, response):
        """内衣"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'内衣'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'内衣')

        for sel in response.xpath("//ul[@class='list-wrap']/../../textarea"):
            dict_old = eval(sel.xpath("text()").extract()[0].strip())
            for idx,c1 in enumerate(dict_old['cat_mian']):
                try:
                    for kw in c1['cat_data']:
                        list_to_write = [idx+1, c1['name'].decode('utf-8'),
                                kw['cat_name'].decode('utf-8')]
#                        print repr(list_to_write).decode('unicode-escape')
                        ws.append(list_to_write)
                except Exception, e:
                    continue
        wb.save(self.path_to_write)


    def crawlingNanzhuang(self, response):
        """男装"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'男装'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'男装')

        for i,sel in enumerate(response.xpath("//dl[@class='theme-bd-level2']")):
            c1 = sel.xpath("dt/div/a/text()").extract()[0]
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
#            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingNvzhuang(self, response):
        """女装"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'女装'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'女装')

        for i,sel in enumerate(response.xpath("//ul[@class='list-wrap']/li")):
            c1 = sel.xpath("p/a/text()").extract()[0]
            list_kws_c1 = sel.xpath("dl/dd/a/text()").extract()
#            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


