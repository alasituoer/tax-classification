#coding:utf-8
import uniout
import PyV8
import json
from openpyxl import Workbook, load_workbook
import time

import sys
reload(sys)
sys.setdefaultencoding('utf-8')

import scrapy

class JDSpider(scrapy.Spider):
    name = 'jd_spider'
    allowed_domains = ['jd.com', '3.cn',]
    path_to_write = '/mnt/hgfs/windows_desktop/classification_and_coding/' +\
            'data/dictionary_building/from_shopping_websites/dict_from_jd_' +\
            time.strftime("%Y%m%d", time.localtime()) + '.xlsx'
   
    def __init__(self):
        try:
            wb = load_workbook(self.path_to_write)
        except Exception, e:
            wb = Workbook()
            wb.save(self.path_to_write)

    def start_requests(self):
        # 抓取首页浮动框全类别及其关键字
        for u in ['https://dc.3.cn/category/get']:
            yield scrapy.Request(url=u, callback=self.crawlingHeadfloat)

        # 抓取首页第一屏左侧访问分类链接后的各子页面
        list_urls = ['https://www.jd.com',]
        for url in list_urls:
            yield scrapy.Request(url=url, callback=self.parse)

    def crawlingHeadfloat(self, response):
        """首页浮动框内的全类别及其关键字"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'首页全类别'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'首页全类别')

        json_data = eval(response.body.decode('gbk'))
        list_old = json_data['data']
        for idx,cates in enumerate(list_old):
            for c1 in cates['s']:
                c1_name = c1['n'].split('|')[1].decode('utf-8')
                for c2 in c1['s']:
                    c2_name = c2['n'].split('|')[1].decode('utf-8')
                    for kw in c2['s']:
                        kw_name = kw['n'].split('|')[1].decode('utf-8')
                        list_to_write = [idx+1, c1_name, c2_name, kw_name,]
#                        print repr(list_to_write).decode('unicode-escape')
                        ws.append(list_to_write)
        wb.save(self.path_to_write)


    def parse(self, response):
        for sel in response.xpath("//ul[@class='JS_navCtn cate_menu']"):
            list_cate_name = sel.xpath("li/a/text()").extract()
            list_cate_url = sel.xpath("li/a/@href").extract()
            list_cate_url = ['https:'+i for i in list_cate_url]
        dict_cate_and_url = dict(zip(list_cate_name, list_cate_url))

        list_cate = [ # 5个一行 共31个品类
                u'家用电器', u'手机', u'数码', u'电脑', u'办公',
                u'家居', u'家具', u'家装', u'厨具', u'男装',
                u'女装', u'童装', u'内衣', u'美妆个护', u'宠物',
                u'女鞋', u'箱包', u'钟表', u'珠宝', u'男鞋',
                u'运动', u'户外', u'汽车', u'汽车用品', u'母婴',
                u'玩具乐器', u'食品', u'酒类', u'生鲜', u'礼品鲜花',
                u'医药保健',]
        list_methods = [
                self.crawlingJiadian, self.crawlingShouji, self.crawlingShuma,
                self.crawlingDiannao, self.crawlingBg, self.crawlingHome,
                self.crawlingFurniture, self.crawlingDecoration,
                self.crawlingKitchenware, self.crawlingMen, self.crawlingWomen,
                self.crawlingChildren, self.crawlingUnderwear, self.crawlingBeauty,
                self.crawlingPet, self.crawlingWomensshoes, self.crawlingBag,
                self.crawlingWatch, self.crawlingJewellery, self.crawlingMensshoes,
                self.crawlingYundongcheng, self.crawlingOutdoor, self.crawlingCar,
                self.crawlingChe, self.crawlingBaby, self.crawlingToy,
                self.crawlingFood, self.crawlingJiu, self.crawlingFresh,
                self.crawlingGiftandFlowers, self.crawlingHealth,]
        dict_cate_and_methods = dict(zip(list_cate, list_methods))

#        for cate_name in [u'手机']:
#        for cate_name in list_cate[-1:]:
        for cate_name in list_cate:
            cate_url = dict_cate_and_url[cate_name]
            #print cate_name, cate_url
            yield scrapy.Request(url=cate_url,
                    callback=dict_cate_and_methods[cate_name])


    def crawlingHealth(self, response):
        """医药保健"""
        body = response.body.decode('gbk', 'ignore')
        start_strings = "window.channelData['floor-firstscreen']"
        end_strings = "})(window);"
        start_index = body.index(start_strings)
        needed_content = body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(' ', '').replace('\n', ''
                ).replace('\t', '')
        needed_content = "var d" + needed_content[len(start_strings):]
#        print needed_content[:100]
#        print needed_content[-100:]

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'医药保健'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'医药保健')

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        for i,c1 in enumerate(dict_content['menu']):
#            print c1['NAME']
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], c2['NAME'],]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingGiftandFlowers(self, response):
        """礼品鲜花"""
        body = response.body.decode('gbk', 'ignore')
        start_strings = "window.channelData['floor-aim-2017110']"
        end_strings = "})(window);"
        start_index = body.index(start_strings)
        needed_content = body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(' ', '').replace('\n', ''
                ).replace('\t', '')
        needed_content = "var d" + needed_content[len(start_strings):]
#        print needed_content[:100]
#        print needed_content[-100:]

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'礼品鲜花'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'礼品鲜花')

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        list_c1 = []
        for i,c1 in enumerate(dict_content['lefttabname']):
#            print c1['NAME']
            list_c1.append(c1['NAME'])
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], '', c2['NAME'],]
#                print list_to_write
                ws.append(list_to_write)

        for i,l in enumerate(dict_content['floatad']):
#            print list_c1[i]
            for c1 in l['floattabcon']:
#                print '\t', c1['NAME']
                for c2 in c1['children']:
#                    print '\t\t', c2['NAME']
                    list_to_write = [i+1, list_c1[i], c1['NAME'], c2['NAME'],]
#                    print list_to_write
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingFresh(self, response):
        """生鲜"""
        body = response.body.decode('gbk', 'ignore')
        start_strings = "navFirst: "
        end_strings = "path:"
        start_index = body.index(start_strings)
        needed_content = body[start_index-len(start_strings)*2:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(' ', '').replace('\n', ''
                ).replace('\t', '')
        needed_content = "var d = {" + needed_content[len(start_strings)+1:-1] + ';'
#        print needed_content[:100]
#        print needed_content[-100:]

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'生鲜'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'生鲜')

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        list_c1 = []
        for i,c1 in enumerate(dict_content['navFirst']):
#            print c1['NAME']
            list_c1.append(c1['NAME'])
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], '', c2['NAME'],]
#                print list_to_write
                ws.append(list_to_write)

        list_section = ['navSecond1', 'navSecond2', 'navSecond3', 'navSecond4',
                'navSecond5',]
        for i,s in enumerate(list_section):
#            print list_c1[i]
            for c1 in dict_content[s]:
#                print '\t', c1['NAME']
                for c2 in c1['children']:
                    list_to_write = [i+1, list_c1[i], c1['NAME'], c2['NAME'],]
#                    print list_to_write
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingJiu(self, response):
        """酒类"""
        start_strings = "window.data['wine_banner_2']"
        # 酒类的结束符与众不同
        end_strings = "isWide: pageConfig.compatible && pageConfig.wideVersion"
        start_index = response.body.index(start_strings)
        needed_content = response.body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(' ', '').replace('\n', ''
                ).replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:] + '};'
#        print needed_content

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'酒类'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'酒类')

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        list_c1 = []
        for i,c1 in enumerate(dict_content['navFirst']):
#            print c1['NAME']
            list_c1.append(c1['NAME'])
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], c2['NAME'],]
#                print list_to_write
                ws.append(list_to_write)

        list_section = ['navThird1', 'navThird2', 'navThird3', 'navThird4',
                'navThird5', 'navThird6', 'navThird7',]
        for i,s in enumerate(list_section):
#            print list_c1[i]
            for c1 in dict_content[s]:
#                print '\t', c1['NAME']
                list_to_write = [i+1, list_c1[i], c1['NAME']]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingFood(self, response):
        """食品"""
        start_strings = "window.data['food_banner_2']"
        end_strings = "})(window);"
        start_index = response.body.index(start_strings)
        needed_content = response.body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(
                'isWide: pageConfig.compatible && pageConfig.wideVersion', ''
                ).replace(' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
#        print needed_content

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'食品'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'食品')

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        list_c1 = []
        for i,c1 in enumerate(dict_content['navFirst']):
#            print c1['NAME']
            list_c1.append(c1['NAME'])
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], '', c2['NAME'],]
#                print list_to_write
                ws.append(list_to_write)

        list_section = ['navThird1', 'navThird2', 'navThird3', 'navThird4',
                'navThird5', 'navThird6',]
        for i,s in enumerate(list_section):
#            print list_c1[i]
            for c1 in dict_content[s]:
#                print '\t', c1['NAME']
                for c2 in c1['children']:
#                    print '\t\t', c2['NAME']
                    list_to_write = [i+1, list_c1[i], c1['NAME'], c2['NAME']]
#                    print list_to_write
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingToy(self, response):
        """玩具乐器"""
        start_strings = "window.data['toys_cms_fs_2']"
        end_strings = "})(window);"
        start_index = response.body.index(start_strings)
        needed_content = response.body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(
                'isWide: pageConfig.compatible && pageConfig.wideVersion', ''
                ).replace(' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
#        print needed_content

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'玩具乐器'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'玩具乐器')

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        for i,c1 in enumerate(dict_content['navFirst']):
#            print c1['NAME']
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], c2['NAME'],]
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingBaby(self, response):
        """母婴"""
        start_strings = "window.data['baby_banner_2']"
        end_strings = "})(window);"
        start_index = response.body.index(start_strings)
        needed_content = response.body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(
                'isWide: pageConfig.compatible && pageConfig.wideVersion', ''
                ).replace(' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
#        print needed_content

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'母婴'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'母婴')

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        list_c1 = []
        for i,c1 in enumerate(dict_content['navFirst']):
#            print c1['NAME']
            list_c1.append(c1['NAME'])
            for d in c1['children']:
#                print '\t', d['NAME']
                list_to_write = [i+1, c1['NAME'], '', d['NAME']]
#                print list_to_write
                ws.append(list_to_write)

        list_section = ['navThird1', 'navThird2', 'navThird3', 'navThird4',
                'navThird5', 'navThird6', 'navThird7', 'navThird8', 'navThird9',]
        for i,s in enumerate(list_section):
#            print list_c1[i]
            for c1 in  dict_content[s]:
#                print '\t', c1['NAME']
                for d in c1['children']:
#                    print '\t\t', d['NAME']
                    list_to_write = [i+1, list_c1[i], c1['NAME'], d['NAME']]
                    ws.append(list_to_write)                    
        wb.save(self.path_to_write)


    def crawlingChe(self, response):
        """汽车用品"""
        start_strings = "window.data['che_banner_2']"
        end_strings = "})(window);"
        start_index = response.body.index(start_strings)
        needed_content = response.body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(
                'isWide: pageConfig.compatible && pageConfig.wideVersion,', ''
                ).replace(' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
#        print needed_content

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'汽车用品'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'汽车用品')

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        list_cate = {6742: u'维修保养' , 6740: u'车载电器', 6743: u'美容清洗',
                    6745: u'汽车装饰', 6747: u'安全自驾', 12402: u'线下服务',
                    13256: u'轮胎配件'}
        for i,s in enumerate(dict_content['menu']):
#            print s['cateId']
            for d in s['category']:
#                print '\t', d['NAME']
                list_to_write = [i+1, list_cate[s['cateId']], d['NAME']]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingCar(self, response):
        """汽车"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'汽车'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'汽车')

        for i,sel in enumerate(response.xpath("//div[@class='categories']/div")):
            c1 = sel.xpath("p/text()").extract()[0].strip()
            list_kws_c1 = sel.xpath("ul/li/a/text()").extract()
#            print c1, list_kws_c1
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingOutdoor(self, response):
        """户外"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'户外'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'户外')

        for i,sel in enumerate(response.xpath("//div[@id='Categorys']/div/dl")):
            c1 = sel.xpath("dt/text()").extract()[0].strip()
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
#            print c1, list_kws_c1
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingYundongcheng(self, response):
        """运动"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'运动'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'运动')

        for i,sel in enumerate(response.xpath("//div[@id='storeCategorys']/div/dl")):
            c1 = sel.xpath("dt/text()").extract()[0].strip()
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
#            print c1, list_kws_c1
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingMensshoes(self, response):
        """男鞋"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'男鞋'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'男鞋')

        for i,sel in enumerate(response.xpath("//div[@class='title']/dl")):
            c1 = sel.xpath("dt/text()").extract()[0].strip()
            list_kws_c1 = sel.xpath("dd/ul/li/a/@title").extract()
#            print c1, list_kws_c1
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingJewellery(self, response):
        """珠宝"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'珠宝'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'珠宝')

        for i,sel in enumerate(response.xpath(
                "//div[@id='jewelleryCategorys']/div/dl")):
            c1 = sel.xpath("dt/text()").extract()[0].strip()
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
#            print c1, list_kws_c1
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingWatch(self, response):
        """钟表"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'钟表'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'钟表')

        list_c1 = []
        for i,sel in enumerate(response.xpath("//div[@class='menu']/div")):
            c1 = sel.xpath("h3/a/text()").extract()[0]
            list_c1.append(c1)
            list_kws_c1 = sel.xpath("div/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw,]
#                print list_to_write
                ws.append(list_to_write)
#        print list_c1

        for i,sel2 in enumerate(response.xpath("//div[@class='sub-menu']/div")):
            list_kws_c1 = sel2.xpath("ul/li/a/@title").extract()
            for kw in list_kws_c1:
                list_to_write = [i+1, list_c1[i], kw,]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingBag(self, response):
        """箱包"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'箱包'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'箱包')

        for i,sel in enumerate(response.xpath(
                "//div[@class='categorys-inner']/div/dl")):
            try:
                c1 = sel.xpath("dt/a/text()").extract()[0].strip()
            except Exception, e:
                c1 = sel.xpath("dt/text()").extract()[0].strip()
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
#            print c1, list_kws_c1
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingWomensshoes(self, response):
        """女鞋"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'女鞋'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'女鞋')

        for i,sel in enumerate(response.xpath("//div[@class='menu']/dl")):
            c1 = sel.xpath("dt/text()").extract()[0].strip()
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
#            print c1, list_kws_c1
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingPet(self, response):
        """宠物"""
        body = response.body#.decode('gbk', 'ignore')
        start_strings = "window.data['pet_slider_2']"
        end_strings = "})(window);"
        start_index = body.index(start_strings)
        needed_content = body[start_index:]
        end_index = needed_content.index(end_strings)

        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(
                'isWide: pageConfig.compatible && pageConfig.wideVersion', ''
                ).replace(' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
        #print needed_content

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'宠物'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'宠物')

        # 根据"navFirst"得到一级分类名列表list_c1及其有限关键字
        list_c1 = []
        for i,c1 in enumerate(dict_content['navFirst']):
#            print c1['NAME']
            list_c1.append(c1['NAME'])
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], '', c2['NAME'],]
#                print list_to_write
                ws.append(list_to_write)
        # 得到二级分类名及其关键字
        list_section = ['navThird1', 'navThird2', 'navThird3', 'navThird4',]
        for i,s in enumerate(list_section):
#            print list_c1[i]
            for c1 in dict_content[s]:
#                print '\t', c1['NAME']
                for c2 in c1['children']:
#                    print '\t\t', c2['NAME']
                    list_to_write = [i+1, list_c1[i], c1['NAME'], c2['NAME'],]
#                    print list_to_write
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingBeauty(self, response):
        """美妆个护"""
        body = response.body.decode('gbk', 'ignore')
        start_strings = "window.data['care_banner2_1']"
        # 美妆个护的结束符有点特别
        end_strings = "};"
        start_index = body.index(start_strings)
        needed_content = body[start_index:]
        end_index = needed_content.index(end_strings)
        # 保证截取字符串完整, 需要加上len(end_index)
        needed_content = needed_content[:end_index+len(end_strings)].strip()
        needed_content = needed_content.replace(
                ' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
#        print needed_content

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'美妆个护'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'美妆个护')

        # 仅提取需要的品类
        list_section = ['subFirst1', 'subFirst2', 'subFirst3', 'subFirst4', 
                'subFirst5', 'subFirst6', 'subFirst7', 'subFirst8',]
        list_section_name = []
        for i,c1 in enumerate(dict_content['navFirst']):
            list_section_name.append(c1['NAME'])
#            print c1['NAME']
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], c2['NAME']]
#                print list_to_write
                ws.append(list_to_write)
#        print list_section_name

        for i,s in enumerate(list_section):
#            print list_section_name[i]
            for c1 in dict_content[s]:
#                print '\t', c1['NAME']
                list_to_write = [i+1, list_section_name[i], c1['NAME']]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingUnderwear(self, response):
        """内衣"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'内衣'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'内衣')

        for i,sel in enumerate(response.xpath("//div[@id='womanCategorys']/div/dl")):
            c1 = sel.xpath("dt/text()").extract()[0].strip()
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
#            print c1, list_kws_c1
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingChildren(self, response):
        """童装"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'童装'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'童装')

        for i,sel in enumerate(response.xpath("//ul[@class='menu']/li")):
            c1 = sel.xpath("h4/text()").extract()[0]
            list_kws_c1 = sel.xpath("div[1]/a/text()").extract()
            c2 = sel.xpath("div[2]/h4/text()").extract()[0]
            list_kws_c2 = sel.xpath("div[2]/div/a/text()").extract()
#            print c1, list_kws_c1
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print list_to_write
                ws.append(list_to_write)

#            print c2, list_kws_c2
            for kw in list_kws_c2:
                list_to_write = [i+1, c2, kw]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingWomen(self, response):
        """女装"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'女装'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'女装')

        list_c1 = []
        for i,sel in enumerate(response.xpath("//div[@class='menu']/dl")):
            c1 = sel.xpath("dt/text()").extract()[0]
            list_c1.append(c1)
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw,]
#                print list_to_write
                ws.append(list_to_write)
#        print list_c1

        for sel2 in response.xpath("//div[@class='sub-menu']/div/div/div"):
            c2 = sel2.xpath("p[1]/text()").extract()[0]
            list_kws_c2 = sel2.xpath("p[2]/a/text()").extract()
#            print c2, list_kws_c2
            for kw in list_kws_c2:
                list_to_write = [list_c1.index(c2)+1, c2, kw,]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingMen(self, response):
        """男装"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'男装'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'男装')

        list_c1 = []
        for i,sel in enumerate(response.xpath("//div[@class='title']/dl")):
            c1 = sel.xpath("dt/text()").extract()[0]
            list_c1.append(c1)
            list_kws_c1 = sel.xpath("dd/ul/li/a/@title").extract()
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, '', kw]
#                print list_to_write
                ws.append(list_to_write)
#        print list_c1

        for i,sel in enumerate(response.xpath("//ul[@class='content']/li/div[1]")):
            for sel2 in sel.xpath("dl"):
                c2 = sel2.xpath("dt/text()").extract()[0]
                list_kws_c2 = sel2.xpath("dd/a/text()").extract()
                for kw in list_kws_c2:
                    list_to_write = [i+1, list_c1[i], c2, kw]
#                    print list_to_write
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingKitchenware(self, response):
        """厨具"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'厨具'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'厨具')

        for i,sel in enumerate(response.xpath("//div[@id='storeCategorys']/div/dl")):
            c1 = sel.xpath("dt/a/text()").extract()[0]
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingDecoration(self, response):
        """家装"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'家装'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'家装')

        for i,sel in enumerate(response.xpath("//div[@id='sortlist']/div/div")):
            c1 = sel.xpath("h3/a/text()").extract()[0]
            list_kws_c1 = sel.xpath("div/a/text()").extract()
#            print c1, list_kws_c1
            # 一级分类及其关键字
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, '', kw]
#                print list_to_write
                ws.append(list_to_write)
            # 二级分类及其关键字
            for sel2 in sel.xpath("div/div[1]/dl"):
                c2 = sel2.xpath("dt/a/text()").extract()[0]
                list_kws_c2 = sel2.xpath("dd/em/a/text()").extract()
                for kw in list_kws_c2:
                    list_to_write = [i+1, c1, c2, kw]
#                    print list_to_write
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingFurniture(self, response):
        """家具"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'家具'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'家具')

        # 得到一级分类名列表list_c1和分类下的关键字list_kws_c1
        list_c1 = []
        for i,sel1 in enumerate(response.xpath("//div[@class='menu']/div/div")):
            c1 = sel1.xpath("div[1]/a/text()").extract()[0]
            list_c1.append(c1)
            list_kws_c1 = sel1.xpath("div[2]/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, '', kw]
#                print list_to_write
                ws.append(list_to_write)
#        print list_c1

        # 得到二级分类名c2及其类别下的关键字列表list_kws_c2
        for i,sel2 in enumerate(response.xpath("//div[@class='sub-menu']/div")):
            for sel3 in sel2.xpath("div/dl"):
                c2 = sel3.xpath("dt/a/text()").extract()[0]
                list_kws_c2 = sel3.xpath("dd/a/text()").extract()
                for kw in list_kws_c2:
                    list_to_write = [i+1, list_c1[i], c2, kw]
#                    print list_to_write
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingHome(self, response):
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'家居'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'家居')

        for i,sel in enumerate(response.xpath("//div[@id='sortlist']/div/div")):
            c1 = sel.xpath("h3/a/text()").extract()[0]
            list_c2 = sel.xpath("div/a/text()").extract()
            list_c3 = sel.xpath("ul/li/a/text()").extract()
#            print c1, list_c2, list_c3
            for c2 in list_c2:
                list_to_write = [i+1, c1, '', c2]
#                print list_to_write
                ws.append(list_to_write)
            for c3 in list_c3:
                list_to_write = [i+1, c1, u'全部分类', c3]
#                print list_to_write
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingBg(self, response):
        """办公"""
        body = response.body#.decode('gbk', 'ignore')
        start_strings = "window.channelData['floor-2']"
        end_strings = "})(window);"
        start_index = body.index(start_strings)
        needed_content = body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(
                ' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
        #print needed_content

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'办公'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'办公')

        # 通过第一个section得到后续一级分类的类名
        list_section_name = []
        for i,c1 in enumerate(dict_content['menu']):
            list_section_name.append(c1['NAME'])
#            print c1['NAME']
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], '', c2['NAME']]
#                print list_to_write
                ws.append(list_to_write)
#        print list_section_name

        for i,l in enumerate(dict_content['submenu']):
#            print list_section_name[i]
            for c1 in l['link']:
#                print '\t', c1['NAME']
                for c2 in c1['children']:
#                    print '\t\t', c2['NAME']
                    list_to_write = [i+1, list_section_name[i], c1['NAME'], c2['NAME'],]
                    ws.append(list_to_write)
#                    print list_to_write
        wb.save(self.path_to_write)


    def crawlingDiannao(self, response):
        """电脑"""
        body = response.body#.decode('gbk', 'ignore')
        start_strings = "window.channelData['floor-2']"
        end_strings = "})(window);"
        start_index = body.index(start_strings)
        needed_content = body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(
                ' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
        #print needed_content

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        wb = load_workbook(self.path_to_write)
        # 覆盖旧表
        try:
            wb.remove_sheet(wb[u'电脑'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'电脑')

        # 通过第一个section得到后续一级分类的类名
        list_section_name = []
        list_to_write = []
        for i,c1 in enumerate(dict_content['menu']):
            list_section_name.append(c1['pic'][0]['NAME'])
#            print c1['pic'][0]['NAME']
            for c2 in c1['link']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['pic'][0]['NAME'], '', c2['NAME']]
                #print list_to_write
                ws.append(list_to_write)
#        print list_section_name

        for i,l in enumerate(dict_content['submenu']):
#            print list_section_name[i]
            for c1 in l['link']:
#                print '\t', c1['NAME']
                for c2 in c1['children']:
#                    print '\t\t', c2['NAME']
                    list_to_write = [i+1, list_section_name[i], c1['NAME'], c2['NAME'],]
                    ws.append(list_to_write)
#                    print list_to_write
        wb.save(self.path_to_write)


    def crawlingShuma(self, response):
        """数码"""
        body = response.body#.decode('gbk', 'ignore')
        start_strings = "window.channelData['floor-3']"
        end_strings = "})(window);"
        start_index = body.index(start_strings)
        needed_content = body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(
                ' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
        #print needed_content

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
#        print dict_content.keys()

        # 通过第一个section得到后续一级分类的类名
        list_section_name = []
        for c1 in dict_content['title']:
            list_section_name.append(c1['NAME'])
        #print list_section_name

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'数码'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'数码')

        for i,l in enumerate(dict_content['content']):
#            print list_section_name[i]
            for c1 in l['link']:
#                print '\t', c1['NAME']
                for c2 in c1['children']:
#                    print '\t\t', c2['NAME']
                    list_to_write = [i+1, list_section_name[i], c1['NAME'], c2['NAME'],]
#                    print list_to_write
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingShouji(self, response):
        """手机"""
        body = response.body.decode('gbk', 'ignore')
        start_strings = "window.channelData['floor-firstscreen']"
        end_strings = "})(window);"
        start_index = body.index(start_strings)
        needed_content = body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(
                ' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
        #print needed_content

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)
        #print dict_content.keys()

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'手机'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'手机')

        # 通过第一个section得到后续一级分类的类名
        list_section_name = []
        for i,c1 in enumerate(dict_content['menu']):
            list_section_name.append(c1['NAME'])
#            print c1['NAME']
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], '', c2['NAME']]
#                print list_to_write
                ws.append(list_to_write)
#        print list_section_name

        for i,l in enumerate(dict_content['submenu']):
#            print list_section_name[i]
            for c1 in l['link']:
#                print '\t', c1['NAME']
                for c2 in c1['children']:
#                    print '\t\t', c2['NAME']
                    list_to_write = [i+1, list_section_name[i], c1['NAME'], c2['NAME'],]
                    ws.append(list_to_write)
#                    print list_to_write
        wb.save(self.path_to_write)


    def crawlingJiadian(self, response):
        """家用电器"""
        start_strings = "window.data['jiadian_banner_2']"
        end_strings = "})(window);"
        start_index = response.body.index(start_strings)
        needed_content = response.body[start_index:]
        end_index = needed_content.index(end_strings)
        needed_content = needed_content[:end_index].strip()
        needed_content = needed_content.replace(
                'isWide: pageConfig.compatible && pageConfig.wideVersion', ''
                ).replace(' ', '').replace('\n', '').replace('\t', '')
        needed_content = "var d = " + needed_content[len(start_strings)+1:]
        #print needed_content[:10]

        pv = PyV8.JSContext()
        pv.enter()
        # 将JS [object Object] 转化为json格式
        pv.eval(needed_content + "var json_data = JSON.stringify(d);")
        # 再将json格式转换为python字典格式
        dict_content = json.loads(pv.locals.json_data)

        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'家电'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'家电')

        # 第一个字典(一级分类)的关键字(品类名)分别是后续多个字典(一级分类)的品类名
        list_section_name = []
        for i,c1 in enumerate(dict_content['navFirst']):
            list_section_name.append(c1['NAME'])
#            print c1['NAME']
            for c2 in c1['children']:
#                print '\t', c2['NAME']
                list_to_write = [i+1, c1['NAME'], '', c2['NAME']]
#                print list_to_write
                ws.append(list_to_write)
#         print list_section_name

        # 仅提取需要的品类
        list_section = ['navThird1', 'navThird2', 'navThird3',
                'navThird4', 'navThird5', 'navThird6', 'navThird7',]
        #print list_section_name

        for i,s in enumerate(list_section):
#            print list_section_name[i]
            for c1 in dict_content[s]:
#                print '\t', c1['NAME']
                for c2 in c1['children']:
#                    print '\t\t', c2['NAME']
                    list_to_write = [i+1, list_section_name[i], c1['NAME'], c2['NAME']]
#                    print list_to_write
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


